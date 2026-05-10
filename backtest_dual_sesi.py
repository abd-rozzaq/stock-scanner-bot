"""
Backtester Dual-Sesi untuk Claude Screener V4.1
================================================
Mensimulasikan screener Sesi 1 + Sesi 2 pada data historis
menggunakan data intraday 5 menit dari Yahoo Finance.

Cara pakai:
  python backtest_dual_sesi.py                    → 3 bulan terakhir
  python backtest_dual_sesi.py --months 1         → 1 bulan terakhir
  python backtest_dual_sesi.py --start 2024-11-01 → dari tanggal tertentu
  python backtest_dual_sesi.py --tickers BBCA TLKM → ticker spesifik

Output:
  backtest_results/backtest_signals.csv           → semua sinyal
  backtest_results/backtest_summary.xlsx          → laporan ringkas
"""

import pandas as pd
import numpy as np
import yfinance as yf
import datetime as dt
import warnings
import argparse
import os
import sys
import time
import logging

warnings.filterwarnings("ignore")

# ── Timezone ──────────────────────────────────────────────────────────────────
try:
    from zoneinfo import ZoneInfo
    WIB = ZoneInfo("Asia/Jakarta")
except ImportError:
    try:
        import pytz
        WIB = pytz.timezone("Asia/Jakarta")
    except ImportError:
        WIB = dt.timezone(dt.timedelta(hours=7))

# ── Output dir ────────────────────────────────────────────────────────────────
OUTPUT_DIR = "backtest_results"
os.makedirs(OUTPUT_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger("BACKTEST")

# ══════════════════════════════════════════════════════════════════════════════
# 1. KONSTANTA (sama persis dengan screener asli)
# ══════════════════════════════════════════════════════════════════════════════
RSI_PERIOD       = 14
MA5_PERIOD       = 5
MA20_PERIOD      = 20
MA50_PERIOD      = 50
CMF_PERIOD       = 14
OBV_SLOPE_PERIOD = 5
AD_SLOPE_PERIOD  = 5

TP_PERCENT       = 0.06
CL_PERCENT       = 0.05
HOLD_DAYS_MAX    = 5      # max hari hold untuk evaluasi TP/CL

BANDAR_VALUE_MA20_MIN = 1_000_000_000
MIN_VALUE_LIKUIDITAS  = 100_000_000

# Jam sesi (WIB)
SESI1_CUT_HOUR, SESI1_CUT_MIN = 11, 30   # bar s/d 11:30 → Sesi 1
SESI2_CUT_HOUR, SESI2_CUT_MIN = 14, 55   # bar s/d 14:55 → Sesi 2

TOTAL_MARKET_MINUTES = 240.0
SESI1_ELAPSED_MIN    = 150.0   # 09:00–11:30
SESI2_ELAPSED_MIN    = 240.0   # seluruh hari

SESSION_PROFILES = {
    "SESI1": {
        "close_high_ratio":     0.975,
        "min_price_change_pct": 2.5,
        "min_frequency":        1_500,
        "rsi_min":              35,  "rsi_max": 75,
        "min_candle_body":      0.25,
        "min_value_idr":        1_500_000_000,
        "min_noise_score":      2,
        "cmf_min":              0.02,
        "max_prerun_5d":        20.0,
        "max_prerun_10d":       38.0,
        "use_projected_vol":    True,
        "vol_spike_mult":       1.5,
        "elapsed_min":          SESI1_ELAPSED_MIN,
    },
    "SESI2": {
        "close_high_ratio":     0.985,
        "min_price_change_pct": 3.0,
        "min_frequency":        2_000,
        "rsi_min":              40,  "rsi_max": 72,
        "min_candle_body":      0.35,
        "min_value_idr":        3_000_000_000,
        "min_noise_score":      3,
        "cmf_min":              0.05,
        "max_prerun_5d":        15.0,
        "max_prerun_10d":       30.0,
        "use_projected_vol":    False,
        "vol_spike_mult":       2.0,
        "elapsed_min":          SESI2_ELAPSED_MIN,
    },
}

SHARED_FILTERS = {
    "min_price_idr":   50,
    "vol_ma20_min":    2.0,
    "max_vol_proj_cap": 3.0,
}

# ══════════════════════════════════════════════════════════════════════════════
# 2. INDIKATOR TEKNIKAL (identik dengan screener asli)
# ══════════════════════════════════════════════════════════════════════════════
def calculate_rsi(series, period=RSI_PERIOD):
    if len(series) < period + 1:
        return -1.0
    delta = series.diff()
    gain  = delta.where(delta > 0, 0.0)
    loss  = (-delta.where(delta < 0, 0.0))
    ag    = gain.ewm(alpha=1.0/period, min_periods=period, adjust=False).mean()
    al    = loss.ewm(alpha=1.0/period, min_periods=period, adjust=False).mean()
    cag, cal = ag.iloc[-1], al.iloc[-1]
    if cal == 0:
        return 100.0 if cag > 0 else 50.0
    return round(100.0 - (100.0 / (1.0 + cag / cal)), 1)


def compute_cmf(df, period=CMF_PERIOD):
    rng = (df["High"] - df["Low"]).replace(0, np.nan)
    mfm = ((df["Close"] - df["Low"]) - (df["High"] - df["Close"])) / rng
    mfm = mfm.fillna(0)
    mfv = mfm * df["Volume"]
    vs  = df["Volume"].rolling(period).sum().replace(0, np.nan)
    cmf = mfv.rolling(period).sum() / vs
    return cmf.fillna(0)


def compute_obv(df):
    direction = np.sign(df["Close"].diff().fillna(0))
    return (direction * df["Volume"]).cumsum()


def compute_ad_line(df):
    rng = (df["High"] - df["Low"]).replace(0, np.nan)
    clv = ((df["Close"] - df["Low"]) - (df["High"] - df["Close"])) / rng
    clv = clv.fillna(0)
    return (clv * df["Volume"]).cumsum()


def get_slope_sign(series, lookback):
    if len(series) < lookback + 1:
        return 0
    recent = series.iloc[-lookback:].values.astype(float)
    if np.any(np.isnan(recent)):
        return 0
    slope = np.polyfit(np.arange(len(recent)), recent, 1)[0]
    return 1 if slope > 0 else (-1 if slope < 0 else 0)

# ══════════════════════════════════════════════════════════════════════════════
# 3. FETCH DATA
# ══════════════════════════════════════════════════════════════════════════════
def fetch_historical(ticker, start_date, end_date):
    """Ambil data harian historis (lebih luas dari window backtest untuk warmup MA)."""
    symbol = f"{ticker}.JK" if not ticker.endswith(".JK") else ticker
    # Ambil 90 hari ekstra sebelum start untuk warmup MA50
    fetch_start = start_date - dt.timedelta(days=120)
    try:
        df = yf.download(
            symbol,
            start=fetch_start.strftime("%Y-%m-%d"),
            end=(end_date + dt.timedelta(days=10)).strftime("%Y-%m-%d"),
            interval="1d",
            progress=False,
            auto_adjust=False,
        )
        if df.empty:
            return pd.DataFrame()
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [c[0] for c in df.columns]
        cols = ["Open", "High", "Low", "Close", "Volume"]
        if not all(c in df.columns for c in cols):
            return pd.DataFrame()
        df.index = pd.to_datetime(df.index).tz_localize(None)
        return df[cols].dropna().copy()
    except Exception as e:
        logger.debug(f"Error fetch daily {ticker}: {e}")
        return pd.DataFrame()


def fetch_intraday(ticker, start_date, end_date):
    """
    Ambil data intraday 5 menit untuk seluruh periode backtest.
    yfinance membatasi 60 hari untuk data 5m, jadi kita fetch per chunk.
    """
    symbol = f"{ticker}.JK" if not ticker.endswith(".JK") else ticker
    all_chunks = []
    chunk_start = start_date
    while chunk_start < end_date:
        chunk_end = min(chunk_start + dt.timedelta(days=55), end_date)
        try:
            df = yf.download(
                symbol,
                start=chunk_start.strftime("%Y-%m-%d"),
                end=(chunk_end + dt.timedelta(days=1)).strftime("%Y-%m-%d"),
                interval="5m",
                progress=False,
                auto_adjust=False,
            )
            if not df.empty:
                if isinstance(df.columns, pd.MultiIndex):
                    df.columns = [c[0] for c in df.columns]
                all_chunks.append(df)
        except Exception as e:
            logger.debug(f"Error fetch intraday chunk {ticker} {chunk_start}: {e}")
        chunk_start = chunk_end + dt.timedelta(days=1)
        time.sleep(0.3)

    if not all_chunks:
        return pd.DataFrame()
    result = pd.concat(all_chunks)
    result = result[~result.index.duplicated(keep='first')]
    result.index = pd.to_datetime(result.index)
    # Normalize ke UTC+7 / naive
    if result.index.tz is not None:
        result.index = result.index.tz_convert("Asia/Jakarta").tz_localize(None)
    return result.sort_index()

# ══════════════════════════════════════════════════════════════════════════════
# 4. CORE ANALYSIS — SESSION-AWARE (BACKTESTING VERSION)
# ══════════════════════════════════════════════════════════════════════════════
def aggregate_intraday_to_session(intra_df, trade_date, session):
    """
    Potong data intraday sesuai sesi:
      SESI1 → 09:00 s/d 11:30
      SESI2 → 09:00 s/d 14:55
    Kembalikan dict OHLCV agregat atau None.
    """
    date_str = trade_date.strftime("%Y-%m-%d")
    day_bars  = intra_df[intra_df.index.date == trade_date.date()]
    if day_bars.empty:
        return None

    if session == "SESI1":
        cut = dt.datetime.combine(trade_date.date(), dt.time(11, 30))
        bars = day_bars[day_bars.index <= cut]
    else:
        cut = dt.datetime.combine(trade_date.date(), dt.time(14, 55))
        bars = day_bars[day_bars.index <= cut]

    bars = bars.dropna(subset=["Close", "Volume"])
    if len(bars) < 3:
        return None

    cols = ["Open", "High", "Low", "Close", "Volume"]
    if not all(c in bars.columns for c in cols):
        return None

    today_open  = float(bars["Open"].iloc[0])
    today_high  = float(bars["High"].max())
    today_low   = float(bars["Low"].min())
    today_close = float(bars["Close"].iloc[-1])
    today_vol   = float(bars["Volume"].sum())

    # Volume projection (Sesi 1)
    cfg = SESSION_PROFILES[session]
    if cfg["use_projected_vol"]:
        elapsed = cfg["elapsed_min"]
        raw_mult = TOTAL_MARKET_MINUTES / elapsed if elapsed > 0 else 1.0
        mult = min(raw_mult, SHARED_FILTERS["max_vol_proj_cap"])
    else:
        mult = 1.0

    return {
        "open":       today_open,
        "high":       today_high,
        "low":        today_low,
        "close":      today_close,
        "volume":     today_vol * mult,
        "volume_raw": today_vol,
        "vol_mult":   round(mult, 2),
    }


def analyze_stock_on_date(ticker, trade_date, session, hist_df, intra_df):
    """
    Analisis satu saham pada tanggal tertentu, mensimulasikan screener asli.
    """
    cfg = SESSION_PROFILES[session]

    # Data harian s/d H-1 (untuk MA, RSI, CMF, anti-pump)
    hist_to_date = hist_df[hist_df.index < pd.Timestamp(trade_date)]
    if len(hist_to_date) < 65:
        return None

    # Intraday hari ini → harga & volume terkini
    intra = aggregate_intraday_to_session(intra_df, trade_date, session)
    if intra is None:
        return None

    close      = intra["close"]
    high       = intra["high"]
    low        = intra["low"]
    open_price = intra["open"]
    volume     = intra["volume"]
    prev_close = float(hist_to_date["Close"].iloc[-1])

    if close <= 0 or high <= 0 or volume <= 0 or prev_close <= 0:
        return None

    # Derived series
    value_series = hist_to_date["Close"] * hist_to_date["Volume"]
    value_today  = close * volume

    vol_ma5    = hist_to_date["Volume"].rolling(MA5_PERIOD).mean().iloc[-1]
    vol_ma20   = hist_to_date["Volume"].rolling(MA20_PERIOD).mean().iloc[-1]
    ma20       = hist_to_date["Close"].rolling(MA20_PERIOD).mean().iloc[-1]
    ma50       = hist_to_date["Close"].rolling(MA50_PERIOD).mean().iloc[-1]
    value_ma20 = value_series.rolling(MA20_PERIOD).mean().iloc[-1]

    if any(pd.isna(x) for x in [vol_ma5, vol_ma20, ma20, ma50, value_ma20]):
        return None

    price_change_pct = ((close - prev_close) / prev_close) * 100

    # Anti-pump
    if len(hist_to_date) < 12:
        return None
    price_5d_ago  = float(hist_to_date["Close"].iloc[-7])
    price_10d_ago = float(hist_to_date["Close"].iloc[-12])
    prior_run_5d  = ((prev_close - price_5d_ago)  / price_5d_ago)  * 100
    prior_run_10d = ((prev_close - price_10d_ago) / price_10d_ago) * 100

    # Indikator bandar
    cmf_series = compute_cmf(hist_to_date)
    obv_series = compute_obv(hist_to_date)
    ad_series  = compute_ad_line(hist_to_date)
    cmf_value  = float(cmf_series.iloc[-1])
    obv_slope  = get_slope_sign(obv_series, OBV_SLOPE_PERIOD)
    ad_slope   = get_slope_sign(ad_series, AD_SLOPE_PERIOD)
    rsi_value  = calculate_rsi(hist_to_date["Close"])

    # Candle quality
    candle_range = high - low
    candle_body  = close - open_price
    body_ratio   = (candle_body / candle_range) if candle_range > 0 else 0.0

    # ── Main Conditions ──────────────────────────────────
    m1 = close >= (high * cfg["close_high_ratio"])
    m2 = volume > cfg["min_frequency"]
    m3 = volume > float(vol_ma5)
    m4 = price_change_pct >= cfg["min_price_change_pct"]
    m5 = (rsi_value >= cfg["rsi_min"]) and (rsi_value <= cfg["rsi_max"])
    m6 = close >= SHARED_FILTERS["min_price_idr"]
    m7 = (candle_body > 0) and (body_ratio >= cfg["min_candle_body"])
    m8 = prior_run_5d  <= cfg["max_prerun_5d"]
    m9 = prior_run_10d <= cfg["max_prerun_10d"]

    if not all([m1, m2, m3, m4, m5, m6, m7, m8, m9]):
        return None

    # ── Noise Filters ────────────────────────────────────
    n1 = value_today >= cfg["min_value_idr"]
    n2 = close > float(ma20)
    n3 = volume >= (cfg["vol_spike_mult"] * float(vol_ma20))
    noise_score = sum([n1, n2, n3])
    if noise_score < cfg["min_noise_score"]:
        return None

    # ── Bandar Screener ──────────────────────────────────
    b1 = cmf_value > cfg["cmf_min"]
    b2 = obv_slope > 0
    b3 = ad_slope  > 0
    b4 = float(value_ma20) > BANDAR_VALUE_MA20_MIN
    if not all([b1, b2, b3, b4]):
        return None

    # ── Trend Screener ───────────────────────────────────
    t1 = close > float(ma20)
    t2 = close > float(ma50)
    t3 = volume >= (SHARED_FILTERS["vol_ma20_min"] * float(vol_ma20))
    t4 = value_today > float(value_ma20)
    if not all([t1, t2, t3, t4]):
        return None

    # ── Likuiditas ───────────────────────────────────────
    l1 = volume > (2 * float(vol_ma20))
    l2 = value_today >= MIN_VALUE_LIKUIDITAS
    if not all([l1, l2]):
        return None

    tp = int(close * (1 + TP_PERCENT))
    cl = int(close * (1 - CL_PERCENT))

    return {
        "ticker":         ticker,
        "date":           trade_date.strftime("%Y-%m-%d"),
        "session":        session,
        "close":          int(close),
        "change_pct":     round(price_change_pct, 2),
        "rsi":            rsi_value,
        "cmf":            round(cmf_value, 4),
        "body_ratio":     round(body_ratio, 3),
        "prior_run_5d":   round(prior_run_5d, 2),
        "prior_run_10d":  round(prior_run_10d, 2),
        "value_b":        round(value_today / 1e9, 3),
        "vol_mult":       intra["vol_mult"],
        "noise_score":    noise_score,
        "entry":          int(close),
        "tp":             tp,
        "cl":             cl,
    }

# ══════════════════════════════════════════════════════════════════════════════
# 5. EVALUASI TP / CL
# ══════════════════════════════════════════════════════════════════════════════
def evaluate_outcome(signal, hist_df, trade_dates_sorted):
    """
    Cek apakah TP atau CL tercapai dalam HOLD_DAYS_MAX hari setelah sinyal.
    Menggunakan daily high/low untuk TP dan daily low untuk CL.
    Urutan evaluasi: CL dievaluasi dulu jika candle sama-sama tembus.
    """
    entry_date  = pd.Timestamp(signal["date"])
    entry_price = signal["entry"]
    tp_price    = signal["tp"]
    cl_price    = signal["cl"]

    # Cari hari trading setelah sinyal (hold_days = jumlah hari TRADING, bukan kalender)
    future_dates = [d for d in trade_dates_sorted if d > entry_date][:HOLD_DAYS_MAX]

    for trading_day_num, fwd_date in enumerate(future_dates, start=1):
        if fwd_date not in hist_df.index:
            continue
        row = hist_df.loc[fwd_date]
        day_high = float(row["High"])
        day_low  = float(row["Low"])

        hit_tp = day_high >= tp_price
        hit_cl = day_low  <= cl_price

        if hit_tp and hit_cl:
            pnl_pct = -CL_PERCENT * 100
            return {"outcome": "CL_AMBIGUOUS", "exit_date": fwd_date.strftime("%Y-%m-%d"),
                    "pnl_pct": round(pnl_pct, 2), "hold_days": trading_day_num}
        elif hit_tp:
            return {"outcome": "TP", "exit_date": fwd_date.strftime("%Y-%m-%d"),
                    "pnl_pct": round(TP_PERCENT * 100, 2), "hold_days": trading_day_num}
        elif hit_cl:
            return {"outcome": "CL", "exit_date": fwd_date.strftime("%Y-%m-%d"),
                    "pnl_pct": round(-CL_PERCENT * 100, 2), "hold_days": trading_day_num}

    # Tidak tembus TP / CL → tutup di harga terakhir available
    last_avail = [(i + 1, d) for i, d in enumerate(future_dates) if d in hist_df.index]
    if last_avail:
        last_tday, last_date = last_avail[-1]
        exit_close = float(hist_df.loc[last_date, "Close"])
        pnl_pct = ((exit_close - entry_price) / entry_price) * 100
        return {"outcome": "EXPIRED", "exit_date": last_date.strftime("%Y-%m-%d"),
                "pnl_pct": round(pnl_pct, 2), "hold_days": last_tday}

    return {"outcome": "NO_DATA", "exit_date": None, "pnl_pct": 0.0, "hold_days": 0}

# ══════════════════════════════════════════════════════════════════════════════
# 6. LOAD TICKERS
# ══════════════════════════════════════════════════════════════════════════════
def load_tickers(tickers_arg=None):
    if tickers_arg:
        return [t.strip().upper() for t in tickers_arg]
    for path in [os.path.join("data", "data.csv"), "data.csv"]:
        if os.path.exists(path):
            df  = pd.read_csv(path)
            col = next(
                (c for c in ["Ticker","ticker","Kode","kode","Code","code","Symbol","symbol"]
                 if c in df.columns),
                df.columns[0]
            )
            tickers = df[col].dropna().astype(str).tolist()
            return sorted(set(t.strip().upper() for t in tickers if len(t.strip()) >= 4))
    raise FileNotFoundError(
        "File data.csv tidak ditemukan. Letakkan di folder 'data/' atau gunakan --tickers."
    )

# ══════════════════════════════════════════════════════════════════════════════
# 7. MAIN BACKTEST LOOP
# ══════════════════════════════════════════════════════════════════════════════
def get_idx_trading_dates(start_date, end_date, sample_ticker="BBCA"):
    """Dapatkan hari-hari IDX buka dengan mengecek data historis BBCA."""
    df = yf.download(
        f"{sample_ticker}.JK",
        start=start_date.strftime("%Y-%m-%d"),
        end=(end_date + dt.timedelta(days=1)).strftime("%Y-%m-%d"),
        interval="1d", progress=False, auto_adjust=False
    )
    if df.empty:
        return []
    df.index = pd.to_datetime(df.index).tz_localize(None)
    return sorted(df.index.tolist())


def run_backtest(tickers, start_date, end_date):
    logger.info(f"Backtest Dual-Sesi | {start_date} s/d {end_date} | {len(tickers)} ticker")
    logger.info("Mengambil hari trading IDX...")
    trade_dates = get_idx_trading_dates(start_date, end_date)
    if not trade_dates:
        logger.error("Tidak bisa ambil hari trading IDX.")
        return []

    logger.info(f"Hari trading: {len(trade_dates)} hari")

    all_signals = []
    total = len(tickers)

    for idx, ticker in enumerate(tickers, 1):
        print(f"\r  [{idx}/{total}] Memproses {ticker:<6}...", end="", flush=True)

        # Fetch data sekali per ticker
        hist_df  = fetch_historical(ticker, start_date, end_date)
        if hist_df.empty or len(hist_df) < 65:
            continue

        intra_df = fetch_intraday(ticker, start_date, end_date)
        if intra_df.empty:
            logger.debug(f"Tidak ada data intraday untuk {ticker}")
            continue

        time.sleep(0.5)  # Rate limit yfinance

        # Loop per hari trading
        for trade_date in trade_dates:
            if trade_date < pd.Timestamp(start_date):
                continue
            if trade_date > pd.Timestamp(end_date):
                break

            # Analisis SESI 1 dan SESI 2
            s1_result = analyze_stock_on_date(ticker, trade_date, "SESI1", hist_df, intra_df)
            s2_result = analyze_stock_on_date(ticker, trade_date, "SESI2", hist_df, intra_df)

            confirmed = (s1_result is not None) and (s2_result is not None)

            if s2_result is not None:
                s2_result["confirmation"] = "CONFIRMED" if confirmed else "NEW"
                # Evaluasi outcome
                outcome = evaluate_outcome(s2_result, hist_df, [d for d in trade_dates])
                s2_result.update(outcome)
                all_signals.append(s2_result)

    print("\r" + " " * 60 + "\r", end="")
    logger.info(f"Selesai. Total sinyal Sesi 2: {len(all_signals)}")
    return all_signals

# ══════════════════════════════════════════════════════════════════════════════
# 8. EKSPOR HASIL
# ══════════════════════════════════════════════════════════════════════════════
def compute_statistics(df):
    stats = {}
    total = len(df)
    if total == 0:
        return stats

    stats["total_sinyal"]    = total
    stats["confirmed"]       = int((df["confirmation"] == "CONFIRMED").sum())
    stats["new"]             = int((df["confirmation"] == "NEW").sum())

    stats["tp_count"]        = int((df["outcome"] == "TP").sum())
    stats["cl_count"]        = int(df["outcome"].isin(["CL", "CL_AMBIGUOUS"]).sum())
    stats["expired_count"]   = int((df["outcome"] == "EXPIRED").sum())

    stats["win_rate_pct"]    = round(stats["tp_count"] / total * 100, 1)
    stats["avg_pnl_pct"]     = round(df["pnl_pct"].mean(), 2)
    stats["median_pnl_pct"]  = round(df["pnl_pct"].median(), 2)
    stats["best_trade_pct"]  = round(df["pnl_pct"].max(), 2)
    stats["worst_trade_pct"] = round(df["pnl_pct"].min(), 2)
    stats["avg_hold_days"]   = round(df["hold_days"].mean(), 1)

    # CONFIRMED only
    conf_df = df[df["confirmation"] == "CONFIRMED"]
    if len(conf_df) > 0:
        stats["conf_win_rate_pct"] = round(
            (conf_df["outcome"] == "TP").sum() / len(conf_df) * 100, 1
        )
        stats["conf_avg_pnl_pct"]  = round(conf_df["pnl_pct"].mean(), 2)
    else:
        stats["conf_win_rate_pct"] = None
        stats["conf_avg_pnl_pct"]  = None

    return stats


def export_to_excel(signals, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl tidak terinstall. Jalankan: pip install openpyxl")
        return

    if not signals:
        logger.warning("Tidak ada sinyal untuk diekspor.")
        return

    df = pd.DataFrame(signals)
    stats = compute_statistics(df)
    wb = Workbook()

    # ── Warna & style ────────────────────────────────────
    HDR_FILL = PatternFill("solid", start_color="1F3864")
    SUB_FILL = PatternFill("solid", start_color="2E75B6")
    CONF_FILL = PatternFill("solid", start_color="FFD700")
    TP_FILL   = PatternFill("solid", start_color="C6EFCE")
    CL_FILL   = PatternFill("solid", start_color="FFC7CE")
    EXP_FILL  = PatternFill("solid", start_color="FFEB9C")
    WHITE     = PatternFill("solid", start_color="FFFFFF")
    LIGHT     = PatternFill("solid", start_color="F2F2F2")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(size=11, bold=True, color="FFFFFF"):
        return Font(name="Arial", bold=bold, size=size, color=color)

    def cell_font(size=10, bold=False, color="000000"):
        return Font(name="Arial", bold=bold, size=size, color=color)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    # ══════════════════════════════════════════════════════
    # SHEET 1: RINGKASAN
    # ══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Ringkasan"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 32
    ws1.column_dimensions["D"].width = 20
    ws1.row_dimensions[1].height = 36
    ws1.row_dimensions[2].height = 22

    # Judul
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value = "📊 LAPORAN BACKTESTING — CLAUDE SCREENER V4.1 DUAL SESI"
    c.font      = hdr_font(13)
    c.fill      = HDR_FILL
    c.alignment = center

    ws1.merge_cells("A2:D2")
    c = ws1["A2"]
    c.value = (
        f"Periode: {df['date'].min()} s/d {df['date'].max()} | "
        f"Ticker: {df['ticker'].nunique()} | TP={TP_PERCENT*100:.0f}%  CL={CL_PERCENT*100:.0f}%  MaxHold={HOLD_DAYS_MAX}H"
    )
    c.font      = hdr_font(10, bold=False, color="DDDDDD")
    c.fill      = HDR_FILL
    c.alignment = center

    # Blok statistik
    sections = [
        ("SEMUA SINYAL SESI 2", [
            ("Total sinyal",          stats.get("total_sinyal"), None),
            ("⭐ CONFIRMED (S1+S2)",   stats.get("confirmed"),   None),
            ("🆕 NEW (hanya S2)",      stats.get("new"),          None),
            ("",                       None,                       None),
            ("✅ TP tercapai",          stats.get("tp_count"),     None),
            ("❌ CL terkena",           stats.get("cl_count"),     None),
            ("⏳ EXPIRED (5H habis)",   stats.get("expired_count"),None),
        ]),
        ("STATISTIK RETURN", [
            ("Win Rate",              stats.get("win_rate_pct"),  "pct"),
            ("Avg P&L",               stats.get("avg_pnl_pct"),   "pct"),
            ("Median P&L",            stats.get("median_pnl_pct"),"pct"),
            ("Best Trade",            stats.get("best_trade_pct"),"pct"),
            ("Worst Trade",           stats.get("worst_trade_pct"),"pct"),
            ("Avg Hold Days",         stats.get("avg_hold_days"), None),
        ]),
        ("CONFIRMED ONLY ⭐", [
            ("Win Rate (Confirmed)",  stats.get("conf_win_rate_pct"), "pct"),
            ("Avg P&L (Confirmed)",   stats.get("conf_avg_pnl_pct"),  "pct"),
        ]),
    ]

    row = 4
    for section_title, items in sections:
        # Sub-header
        ws1.merge_cells(f"A{row}:D{row}")
        c = ws1[f"A{row}"]
        c.value     = section_title
        c.font      = hdr_font(11)
        c.fill      = SUB_FILL
        c.alignment = left
        c.border    = border
        ws1.row_dimensions[row].height = 20
        row += 1

        for label, val, fmt in items:
            if not label:
                row += 1
                continue
            c_label = ws1.cell(row=row, column=1, value=label)
            c_label.font      = cell_font(10, bold=False)
            c_label.alignment = left
            c_label.border    = border
            c_label.fill      = LIGHT if row % 2 == 0 else WHITE

            display_val = ""
            if val is not None:
                if fmt == "pct":
                    display_val = f"{val:.1f}%"
                    c_val = ws1.cell(row=row, column=2, value=val/100)
                    c_val.number_format = "0.0%"
                else:
                    display_val = val
                    c_val = ws1.cell(row=row, column=2, value=val)
                    c_val.number_format = "#,##0.0"
            else:
                c_val = ws1.cell(row=row, column=2, value="N/A")

            c_val.font      = cell_font(10, bold=True)
            c_val.alignment = right
            c_val.border    = border
            c_val.fill      = LIGHT if row % 2 == 0 else WHITE

            # Warna khusus untuk metrik penting
            if "Win Rate" in label and val is not None:
                color = "00B050" if val >= 50 else "FF0000"
                c_val.font = Font(name="Arial", bold=True, size=11, color=color)

            ws1.row_dimensions[row].height = 18
            row += 1
        row += 1

    # ══════════════════════════════════════════════════════
    # SHEET 2: DETAIL SINYAL
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detail Sinyal")
    ws2.sheet_view.showGridLines = False

    columns = [
        ("Tanggal",        "date",          14),
        ("Ticker",         "ticker",        10),
        ("Sesi",           "session",        8),
        ("Konfirmasi",     "confirmation",  14),
        ("Close",          "close",         10),
        ("Chg%",           "change_pct",    10),
        ("RSI",            "rsi",            8),
        ("CMF",            "cmf",            8),
        ("Body",           "body_ratio",     8),
        ("5dRun%",         "prior_run_5d",  10),
        ("Val(B)",         "value_b",        9),
        ("Noise",          "noise_score",    8),
        ("Entry",          "entry",          9),
        ("TP",             "tp",             9),
        ("CL",             "cl",             9),
        ("Outcome",        "outcome",       14),
        ("Exit Date",      "exit_date",     12),
        ("P&L%",           "pnl_pct",       10),
        ("Hold Days",      "hold_days",      9),
    ]

    # Header
    ws2.row_dimensions[1].height = 24
    for col_i, (header, _, width) in enumerate(columns, 1):
        c = ws2.cell(row=1, column=col_i, value=header)
        c.font      = hdr_font(10)
        c.fill      = HDR_FILL
        c.alignment = center
        c.border    = border
        ws2.column_dimensions[get_column_letter(col_i)].width = width

    # Data rows
    for row_i, sig in enumerate(signals, 2):
        ws2.row_dimensions[row_i].height = 16

        outcome   = sig.get("outcome", "")
        conf      = sig.get("confirmation", "")
        row_fill  = (
            CONF_FILL if conf == "CONFIRMED" and outcome == "TP" else
            TP_FILL   if outcome == "TP" else
            CL_FILL   if outcome in ["CL", "CL_AMBIGUOUS"] else
            EXP_FILL  if outcome == "EXPIRED" else
            WHITE
        )

        for col_i, (_, field, _) in enumerate(columns, 1):
            val = sig.get(field)
            c   = ws2.cell(row=row_i, column=col_i, value=val)
            c.font      = cell_font(9)
            c.fill      = row_fill
            c.border    = border
            c.alignment = center

            if field == "pnl_pct" and val is not None:
                color = "00B050" if float(val) > 0 else ("FF0000" if float(val) < 0 else "000000")
                c.font = Font(name="Arial", size=9, color=color, bold=True)
            if field == "confirmation":
                if val == "CONFIRMED":
                    c.font = Font(name="Arial", size=9, bold=True, color="7B3F00")
                elif val == "NEW":
                    c.font = Font(name="Arial", size=9, color="0070C0")

    # Auto-filter
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws2.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════
    # SHEET 3: STATISTIK PER TICKER
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Per Ticker")
    ws3.sheet_view.showGridLines = False

    ticker_stats = (
        df.groupby("ticker")
        .agg(
            sinyal=("ticker", "count"),
            confirmed=("confirmation", lambda x: (x == "CONFIRMED").sum()),
            tp=("outcome", lambda x: (x == "TP").sum()),
            cl=("outcome", lambda x: x.isin(["CL", "CL_AMBIGUOUS"]).sum()),
            avg_pnl=("pnl_pct", "mean"),
            best=("pnl_pct", "max"),
            worst=("pnl_pct", "min"),
        )
        .reset_index()
    )
    ticker_stats["win_rate"] = (ticker_stats["tp"] / ticker_stats["sinyal"] * 100).round(1)
    ticker_stats["avg_pnl"]  = ticker_stats["avg_pnl"].round(2)
    ticker_stats = ticker_stats.sort_values("avg_pnl", ascending=False)

    t3_cols = [
        ("Ticker",      "ticker",    12),
        ("Sinyal",      "sinyal",    10),
        ("Confirmed",   "confirmed", 12),
        ("TP",          "tp",         8),
        ("CL",          "cl",         8),
        ("Win Rate%",   "win_rate",  12),
        ("Avg P&L%",    "avg_pnl",  12),
        ("Best%",       "best",      10),
        ("Worst%",      "worst",     10),
    ]
    ws3.row_dimensions[1].height = 22
    for ci, (hdr, _, w) in enumerate(t3_cols, 1):
        c = ws3.cell(row=1, column=ci, value=hdr)
        c.font = hdr_font(10)
        c.fill = HDR_FILL
        c.alignment = center
        c.border = border
        ws3.column_dimensions[get_column_letter(ci)].width = w

    for ri, row_data in enumerate(ticker_stats.itertuples(), 2):
        ws3.row_dimensions[ri].height = 16
        fill = LIGHT if ri % 2 == 0 else WHITE
        for ci, (_, field, _) in enumerate(t3_cols, 1):
            val = getattr(row_data, field, None)
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font = cell_font(9)
            c.fill = fill
            c.alignment = center
            c.border = border
            if field == "avg_pnl" and val is not None:
                color = "00B050" if float(val) > 0 else ("FF0000" if float(val) < 0 else "000000")
                c.font = Font(name="Arial", size=9, color=color, bold=True)

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(t3_cols))}1"
    ws3.freeze_panes = "A2"

    # ── Simpan ───────────────────────────────────────────
    wb.save(output_path)
    logger.info(f"Excel tersimpan → {output_path}")


def export_to_csv(signals, output_path):
    if not signals:
        logger.warning("Tidak ada sinyal untuk diekspor ke CSV.")
        return
    df = pd.DataFrame(signals)
    df.to_csv(output_path, index=False, encoding="utf-8-sig")
    logger.info(f"CSV tersimpan → {output_path}")

# ══════════════════════════════════════════════════════════════════════════════
# 9. ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
def parse_args():
    parser = argparse.ArgumentParser(description="Backtester Dual-Sesi Claude Screener")
    parser.add_argument("--months",  type=int,   default=3,
                        help="Jumlah bulan ke belakang (default: 3)")
    parser.add_argument("--start",   type=str,   default=None,
                        help="Tanggal mulai YYYY-MM-DD (override --months)")
    parser.add_argument("--end",     type=str,   default=None,
                        help="Tanggal akhir YYYY-MM-DD (default: hari ini)")
    parser.add_argument("--tickers", nargs="+",  default=None,
                        help="Daftar ticker manual, misal: BBCA TLKM BBRI")
    return parser.parse_args()


def main():
    args = parse_args()

    today = dt.date.today()
    end_date   = dt.datetime.strptime(args.end, "%Y-%m-%d").date() if args.end else today
    if args.start:
        start_date = dt.datetime.strptime(args.start, "%Y-%m-%d").date()
    else:
        # Hitung N bulan ke belakang
        month  = end_date.month - args.months
        year   = end_date.year + (month - 1) // 12
        month  = (month - 1) % 12 + 1
        start_date = end_date.replace(year=year, month=month)

    tickers = load_tickers(args.tickers)

    print()
    print("=" * 70)
    print("  BACKTESTER DUAL-SESI — Claude Screener V4.1")
    print(f"  Periode  : {start_date} s/d {end_date}")
    print(f"  Ticker   : {len(tickers)} saham")
    print(f"  TP/CL    : {TP_PERCENT*100:.0f}% / {CL_PERCENT*100:.0f}%  |  Max Hold: {HOLD_DAYS_MAX} hari")
    print("=" * 70)
    print()

    signals = run_backtest(tickers, start_date, end_date)

    if not signals:
        print("\n⚠️  Tidak ada sinyal ditemukan dalam periode ini.")
        return

    # Export
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path   = os.path.join(OUTPUT_DIR, f"backtest_signals_{ts}.csv")
    xlsx_path  = os.path.join(OUTPUT_DIR, f"backtest_summary_{ts}.xlsx")

    export_to_csv(signals, csv_path)
    export_to_excel(signals, xlsx_path)

    # Print ringkasan ke terminal
    df    = pd.DataFrame(signals)
    stats = compute_statistics(df)

    print()
    print("=" * 70)
    print("  HASIL BACKTESTING")
    print("=" * 70)
    print(f"  Total sinyal Sesi 2  : {stats['total_sinyal']}")
    print(f"  ⭐ CONFIRMED (S1+S2)  : {stats['confirmed']}")
    print(f"  🆕 NEW (hanya S2)     : {stats['new']}")
    print()
    print(f"  ✅ TP tercapai        : {stats['tp_count']}")
    print(f"  ❌ CL terkena         : {stats['cl_count']}")
    print(f"  ⏳ EXPIRED            : {stats['expired_count']}")
    print()
    print(f"  Win Rate (all)        : {stats['win_rate_pct']}%")
    print(f"  Avg P&L               : {stats['avg_pnl_pct']}%")
    print(f"  Best trade            : {stats['best_trade_pct']}%")
    print(f"  Worst trade           : {stats['worst_trade_pct']}%")
    print()
    if stats["conf_win_rate_pct"] is not None:
        print(f"  Win Rate (CONFIRMED)  : {stats['conf_win_rate_pct']}%")
        print(f"  Avg P&L (CONFIRMED)   : {stats['conf_avg_pnl_pct']}%")
    print("=" * 70)
    print(f"\n  📄 CSV  : {csv_path}")
    print(f"  📊 Excel: {xlsx_path}")
    print()


if __name__ == "__main__":
    main()